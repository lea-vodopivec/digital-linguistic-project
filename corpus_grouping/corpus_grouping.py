import openpyxl
import pandas as pd
import numpy as np

def clean_excel(file_path, cleaned_path):
    wb = openpyxl.load_workbook(file_path) # Load the Excel workbook
    ws = wb.active

    rows_to_keep = [] # List to hold rows that are not "idiom" rows
    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            rows_to_keep.append(row)
        else:
            if str(row[0]).strip().lower() != "idiom": # Keep rows where the first cell is not "idiom"
                rows_to_keep.append(row)

    # Remove all data from worksheet before rewriting
    ws.delete_cols(1, ws.max_column)
    ws.delete_rows(1, ws.max_row)

    # Write the kept rows back, excluding the first column
    for r_idx, row in enumerate(rows_to_keep, start=1):
        for c_idx, cell_value in enumerate(row[1:], start=1):
            ws.cell(row=r_idx, column=c_idx, value=cell_value)

    # Save the cleaned file
    wb.save(cleaned_path)
    print(f"Cleaned file saved as: {cleaned_path}")

def regroup_and_save_with_blanks(cleaned_path, output_txt):
    df = pd.read_excel(cleaned_path, header=None) # Load the cleaned data into pandas DataFrame
    sentences = df.values.flatten() # Flatten the data into a list of sentences
    sentences = [s for s in sentences if pd.notna(s)]

    # Shuffle sentences to mix original groups
    np.random.seed(42) # For reproducible shuffles
    np.random.shuffle(sentences)

    groups = [sentences[i:i+10] for i in range(0, len(sentences), 10)] # Divide sentences into groups of 10

    # Write each group into a text file with blank lines between groups
    with open(output_txt, "w", encoding="utf-8") as f:
        for group in groups:
            for sentence in group:
                f.write(sentence + "\n")
            f.write("\n")

    print(f"New grouped sentences saved as text file with blank lines: {output_txt}")

if __name__ == "__main__":
    input_file = "Corpus_of_idioms.xlsx"  # Original file path
    cleaned_file = "Cleaned_Corpus_of_idioms.xlsx"  # Path for cleaned data
    output_file = "ReGrouped_Corpus_of_idioms.txt"  # Output text file

    # Run cleaning procedure
    clean_excel(input_file, cleaned_file)
    # Run regrouping and output with blank lines
    regroup_and_save_with_blanks(cleaned_file, output_file)